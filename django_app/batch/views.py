import io
import logging

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, generics
from rest_framework.permissions import IsAuthenticatedOrReadOnly, IsAuthenticated

from openpyxl import load_workbook
from django.db import transaction

from .models import BatchUpload, BatchItem
from .serializers import BatchUploadSerializer, BatchUploadCreateSerializer, BatchItemSerializer
from .tasks import process_batch_item

from rest_framework.exceptions import PermissionDenied
from django.shortcuts import get_object_or_404
from django.utils.dateparse import parse_datetime
from decimal import Decimal

from .pagination import StandardResultsSetPagination


logger = logging.getLogger(__name__)


class BatchUploadCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, format=None):
        serializer = BatchUploadCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        file_obj = serializer.validated_data['file']

        # Create the BatchUpload record
        batch = BatchUpload.objects.create(
            original_filename=getattr(file_obj, 'name', ''),
            file=file_obj,
            status=BatchUpload.STATUS_PROCESSING,
            uploaded_by=request.user,
        )

        # Parse Excel and create items
        try:
            in_memory = file_obj.read()
            wb = load_workbook(filename=io.BytesIO(in_memory), data_only=True)
            sheet = wb.active

            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                raise ValueError('Uploaded file is empty')

            # Expect header on first row with 'phone' and 'amount' or assume first two columns
            header = [str(h).strip().lower() if h is not None else '' for h in rows[0]]
            has_header = 'phone' in header and 'amount' in header

            data_rows = rows[1:] if has_header else rows

            created_items = []
            with transaction.atomic():
                for i, row in enumerate(data_rows, start=1):
                    if not row or row[0] is None:
                        continue
                    if has_header:
                        phone = row[header.index('phone')]
                        amount = row[header.index('amount')]
                    else:
                        phone = row[0]
                        amount = row[1] if len(row) > 1 else None

                    if phone is None or amount is None:
                        # skip invalid rows
                        continue

                    item = BatchItem.objects.create(
                        batch=batch,
                        row_number=i,
                        phone=str(phone).strip(),
                        amount=round(float(amount), 2),
                    )
                    created_items.append(item)

            batch.total_rows = len(created_items)
            batch.save(update_fields=['total_rows'])

            # Enqueue tasks for each item (auto-start)
            for item in created_items:
                process_batch_item.apply_async(args=(item.id,))

            response = BatchUploadSerializer(batch)
            return Response(response.data, status=status.HTTP_201_CREATED)

        except Exception as exc:
            logger.exception('Failed to parse or process uploaded batch')
            batch.status = BatchUpload.STATUS_FAILED
            batch.save(update_fields=['status'])
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)


class BatchUploadDetailView(generics.RetrieveAPIView):
    queryset = BatchUpload.objects.all()
    serializer_class = BatchUploadSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]

class BatchItemListView(generics.ListAPIView):
    """List items for a given batch with pagination and basic filtering.

    Supported query params:
      - status: exact match (pending, success, failed, ...)
      - phone: substring match (icontains)
      - row: exact row number
      - min_row, max_row
      - min_amount, max_amount
      - processed_before, processed_after (ISO datetime)
      - ordering: comma-separated fields (prefix with - for desc)
    """
    serializer_class = BatchItemSerializer
    pagination_class = StandardResultsSetPagination
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        ALLOWED_ORDERING = [
            'id', 'row_number', 'phone', 'amount', 'status', 'processed_at'
        ]
        batch_id = self.kwargs.get('batch_id')
        # ensure batch exists and enforce simple ownership rule: only uploader or staff can view
        batch = get_object_or_404(BatchUpload, pk=batch_id)
        user = self.request.user
        if batch.uploaded_by and batch.uploaded_by != user and not user.is_staff:
            raise PermissionDenied('You do not have permission to view items for this batch')

        qs = BatchItem.objects.filter(batch_id=batch_id)
        params = self.request.query_params

        status = params.get('status')
        if status:
            qs = qs.filter(status__iexact=status)

        phone = params.get('phone')
        if phone:
            qs = qs.filter(phone__icontains=phone)

        row = params.get('row')
        if row:
            try:
                qs = qs.filter(row_number=int(row))
            except ValueError:
                pass

        min_row = params.get('min_row')
        if min_row:
            try:
                qs = qs.filter(row_number__gte=int(min_row))
            except ValueError:
                pass

        max_row = params.get('max_row')
        if max_row:
            try:
                qs = qs.filter(row_number__lte=int(max_row))
            except ValueError:
                pass

        min_amount = params.get('min_amount')
        if min_amount:
            try:
                qs = qs.filter(amount__gte=Decimal(min_amount))
            except ValueError:
                pass

        max_amount = params.get('max_amount')
        if max_amount:
            try:
                qs = qs.filter(amount__lte=Decimal(max_amount))
            except ValueError:
                pass

        processed_before = params.get('processed_before')
        processed_after = params.get('processed_after')

        if processed_before:
            dt = parse_datetime(processed_before)
            if dt:
                qs = qs.filter(processed_at__lte=dt)

        if processed_after:
            dt = parse_datetime(processed_after)
            if dt:
                qs = qs.filter(processed_at__gte=dt)

        ordering = params.get('ordering')
        if ordering:
            fields = []
            for f in ordering.split(','):
                f = f.strip()
                if f.lstrip('-') in ALLOWED_ORDERING:
                    fields.append(f)
            if fields:
                qs = qs.order_by(*fields)

        return qs
